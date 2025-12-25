"""
Generate test XLSX files for credit application testing
- Income Consistency: Bank statements with 3 months income
- Loan Payments: History of loan payments
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random


def create_income_consistency_xlsx(filename="income_consistency_3months.xlsx"):
    """Create bank statement showing 3 months of income"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Bank header
    ws['A1'] = "ATTIJARIWAFA BANK"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Relevé de Compte / Bank Statement"
    ws['A2'].font = Font(size=12, bold=True)
    ws.merge_cells('A2:F2')
    
    ws['A3'] = f"Période: {(datetime.now() - timedelta(days=90)).strftime('%d/%m/%Y')} - {datetime.now().strftime('%d/%m/%Y')}"
    ws.merge_cells('A3:F3')
    
    ws['A4'] = "Titulaire: Ahmed Benali"
    ws['A5'] = "N° Compte: 007 780 0123456789 12"
    ws['A6'] = ""
    
    # Column headers
    headers = ['Date', 'Description', 'Référence', 'Débit (MAD)', 'Crédit (MAD)', 'Solde (MAD)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Generate 3 months of transactions
    row = 8
    balance = 15000.00  # Starting balance
    current_date = datetime.now() - timedelta(days=90)
    
    # Monthly salary entries
    salaries = [
        ("Virement Salaire - Entreprise ABC", 12000.00),
        ("Virement Salaire - Entreprise ABC", 12000.00),
        ("Virement Salaire - Entreprise ABC", 12000.00),
    ]
    
    # Generate transactions
    transactions = []
    
    # Month 1
    salary_date_1 = current_date + timedelta(days=5)
    transactions.append((salary_date_1, salaries[0][0], f"VIR{random.randint(100000, 999999)}", 0, salaries[0][1]))
    
    # Add some expenses in month 1
    for i in range(8):
        expense_date = salary_date_1 + timedelta(days=random.randint(1, 25))
        amount = random.uniform(200, 1500)
        descriptions = ["Carte Bancaire", "Prélèvement Loyer", "Facture Eau/Elec", "Retrait GAB", "Achat Carrefour"]
        transactions.append((expense_date, random.choice(descriptions), f"REF{random.randint(10000, 99999)}", amount, 0))
    
    # Month 2
    salary_date_2 = current_date + timedelta(days=35)
    transactions.append((salary_date_2, salaries[1][0], f"VIR{random.randint(100000, 999999)}", 0, salaries[1][1]))
    
    # Add expenses in month 2
    for i in range(10):
        expense_date = salary_date_2 + timedelta(days=random.randint(1, 25))
        amount = random.uniform(150, 2000)
        descriptions = ["Carte Bancaire", "Prélèvement Assurance", "Facture Téléphone", "Retrait GAB", "Achat Marjane"]
        transactions.append((expense_date, random.choice(descriptions), f"REF{random.randint(10000, 99999)}", amount, 0))
    
    # Month 3
    salary_date_3 = current_date + timedelta(days=65)
    transactions.append((salary_date_3, salaries[2][0], f"VIR{random.randint(100000, 999999)}", 0, salaries[2][1]))
    
    # Add expenses in month 3
    for i in range(7):
        expense_date = salary_date_3 + timedelta(days=random.randint(1, 20))
        amount = random.uniform(100, 1800)
        descriptions = ["Carte Bancaire", "Virement", "Facture Internet", "Retrait GAB", "Achat Aswak Assalam"]
        transactions.append((expense_date, random.choice(descriptions), f"REF{random.randint(10000, 99999)}", amount, 0))
    
    # Sort by date
    transactions.sort(key=lambda x: x[0])
    
    # Write transactions
    for trans in transactions:
        date, desc, ref, debit, credit = trans
        
        if credit > 0:
            balance += credit
        else:
            balance -= debit
        
        ws.cell(row=row, column=1).value = date.strftime('%d/%m/%Y')
        ws.cell(row=row, column=2).value = desc
        ws.cell(row=row, column=3).value = ref
        ws.cell(row=row, column=4).value = f"{debit:.2f}" if debit > 0 else ""
        ws.cell(row=row, column=5).value = f"{credit:.2f}" if credit > 0 else ""
        ws.cell(row=row, column=6).value = f"{balance:.2f}"
        
        # Apply borders
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            if col >= 4:  # Right align numbers
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        
        # Highlight salary rows
        if credit > 0:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="E7F3E7", end_color="E7F3E7", fill_type="solid")
        
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=1).value = "RÉSUMÉ / SUMMARY"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws.cell(row=row, column=1).value = "Total Crédits (3 mois):"
    ws.cell(row=row, column=1).font = Font(bold=True)
    total_credits = sum(s[1] for s in salaries)
    ws.cell(row=row, column=2).value = f"{total_credits:.2f} MAD"
    ws.cell(row=row, column=2).font = Font(bold=True, color="008000")
    
    row += 1
    ws.cell(row=row, column=1).value = "Revenus Mensuels Moyens:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = f"{total_credits / 3:.2f} MAD"
    ws.cell(row=row, column=2).font = Font(bold=True, color="008000")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    wb.save(filename)
    print(f"✅ Created {filename}")
    print(f"   - 3 months of salary deposits (12,000 MAD each)")
    print(f"   - Multiple expense transactions")
    print(f"   - Total income: 36,000 MAD")


def create_loan_payments_xlsx(filename="loan_payment_history.xlsx"):
    """Create loan payment history spreadsheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan History"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Bank header
    ws['A1'] = "ATTIJARIWAFA BANK - CRÉDIT IMMOBILIER"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws.merge_cells('A1:G1')
    
    ws['A2'] = "Historique des Paiements / Payment History"
    ws['A2'].font = Font(size=12, bold=True)
    ws.merge_cells('A2:G2')
    
    ws['A3'] = "Client: Ahmed Benali"
    ws['A4'] = "N° Crédit: IMM-2020-456789"
    ws['A5'] = f"Montant Initial: 500,000.00 MAD"
    ws['A6'] = f"Durée: 20 ans (240 mois)"
    ws['A7'] = f"Taux: 4.25% annuel"
    ws['A8'] = ""
    
    # Column headers
    headers = ['Mois', 'Date Paiement', 'Mensualité (MAD)', 'Principal (MAD)', 'Intérêts (MAD)', 'Statut', 'Solde Restant (MAD)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Generate 24 months of payments (2 years)
    loan_amount = 500000.00
    annual_rate = 0.0425
    monthly_rate = annual_rate / 12
    months = 240
    monthly_payment = loan_amount * (monthly_rate * (1 + monthly_rate)**months) / ((1 + monthly_rate)**months - 1)
    
    remaining_balance = loan_amount
    row = 10
    start_date = datetime.now() - timedelta(days=730)  # 2 years ago
    
    for month in range(1, 25):
        payment_date = start_date + timedelta(days=30*month)
        
        # Calculate interest and principal
        interest = remaining_balance * monthly_rate
        principal = monthly_payment - interest
        remaining_balance -= principal
        
        # Random payment status (most on time)
        statuses = ["Payé à temps"] * 20 + ["Payé (2j retard)"] * 3 + ["Payé à temps"]
        status = random.choice(statuses)
        
        ws.cell(row=row, column=1).value = f"Mois {month}"
        ws.cell(row=row, column=2).value = payment_date.strftime('%d/%m/%Y')
        ws.cell(row=row, column=3).value = f"{monthly_payment:.2f}"
        ws.cell(row=row, column=4).value = f"{principal:.2f}"
        ws.cell(row=row, column=5).value = f"{interest:.2f}"
        ws.cell(row=row, column=6).value = status
        ws.cell(row=row, column=7).value = f"{remaining_balance:.2f}"
        
        # Apply borders and alignment
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
            if col >= 3:  # Right align numbers
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        
        # Color code status
        if "retard" in status:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        else:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=1).value = "RÉSUMÉ"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    ws.cell(row=row, column=1).value = "Paiements effectués:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = "24 / 24"
    ws.cell(row=row, column=2).font = Font(bold=True, color="008000")
    
    row += 1
    ws.cell(row=row, column=1).value = "Retards:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = "1 paiement (2 jours)"
    ws.cell(row=row, column=2).font = Font(bold=True, color="FF8C00")
    
    row += 1
    ws.cell(row=row, column=1).value = "Total payé:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = f"{monthly_payment * 24:.2f} MAD"
    ws.cell(row=row, column=2).font = Font(bold=True)
    
    row += 1
    ws.cell(row=row, column=1).value = "Statut:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = "Compte en règle ✓"
    ws.cell(row=row, column=2).font = Font(bold=True, color="008000")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    
    wb.save(filename)
    print(f"✅ Created {filename}")
    print(f"   - 24 months of loan payments")
    print(f"   - Monthly payment: {monthly_payment:.2f} MAD")
    print(f"   - Excellent payment history")


if __name__ == "__main__":
    print("📊 Generating test XLSX files...\n")
    
    create_income_consistency_xlsx()
    print()
    create_loan_payments_xlsx()
    
    print("\n✅ All XLSX files generated successfully!")
    print("\n📁 Files created in current directory:")
    print("   1. income_consistency_3months.xlsx")
    print("   2. loan_payment_history.xlsx")
    print("\n💡 Use these files to test the document upload feature")
