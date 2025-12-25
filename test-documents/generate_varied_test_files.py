"""
Generate varied test documents with different financial profiles
to produce different credit scores from ML models
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from datetime import datetime, timedelta
import random


# ===== EXCEL FILES =====

def create_excellent_income_xlsx(filename="excellent_income_consistency.xlsx"):
    """High salary, consistent for 6 months - Score: 95-100"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Header
    ws['A1'] = "ATTIJARIWAFA BANK - EXCELLENT PROFILE"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws.merge_cells('A1:F1')
    ws['A2'] = "Client: Fatima Zahra - Senior Manager"
    ws['A3'] = f"Période: 6 mois - {(datetime.now() - timedelta(days=180)).strftime('%d/%m/%Y')} - {datetime.now().strftime('%d/%m/%Y')}"
    ws.merge_cells('A3:F3')
    ws['A4'] = ""
    
    # Column headers
    headers = ['Date', 'Description', 'Référence', 'Débit (MAD)', 'Crédit (MAD)', 'Solde (MAD)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Generate 6 months of HIGH consistent salary
    row = 6
    balance = 35000.00
    monthly_salary = 18000.00  # High salary
    
    for month in range(6):
        salary_date = datetime.now() - timedelta(days=150 - month*30)
        balance += monthly_salary
        
        ws.cell(row=row, column=1).value = salary_date.strftime('%d/%m/%Y')
        ws.cell(row=row, column=2).value = f"Virement Salaire - Mois {month+1}"
        ws.cell(row=row, column=3).value = f"VIR{random.randint(100000, 999999)}"
        ws.cell(row=row, column=4).value = ""
        ws.cell(row=row, column=5).value = f"{monthly_salary:.2f}"
        ws.cell(row=row, column=6).value = f"{balance:.2f}"
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E7F3E7", end_color="E7F3E7", fill_type="solid")
        
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=1).value = "Total Revenus (6 mois):"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = f"{monthly_salary * 6:.2f} MAD"
    ws.cell(row=row, column=2).font = Font(bold=True, color="008000")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15 if col in ['D', 'E', 'F'] else 20
    
    wb.save(filename)
    print(f"✅ {filename}: Excellent (18,000 MAD x 6 months)")


def create_good_income_xlsx(filename="good_income_consistency.xlsx"):
    """Average salary, consistent - Score: 75-85"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = "ATTIJARIWAFA BANK - GOOD PROFILE"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws.merge_cells('A1:F1')
    ws['A2'] = "Client: Youssef Alami - Employee"
    ws['A3'] = f"Période: 3 mois"
    ws['A4'] = ""
    
    headers = ['Date', 'Description', 'Référence', 'Débit (MAD)', 'Crédit (MAD)', 'Solde (MAD)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 6
    balance = 8000.00
    monthly_salary = 8500.00  # Average salary
    
    for month in range(3):
        salary_date = datetime.now() - timedelta(days=75 - month*25)
        balance += monthly_salary
        
        ws.cell(row=row, column=1).value = salary_date.strftime('%d/%m/%Y')
        ws.cell(row=row, column=2).value = "Virement Salaire"
        ws.cell(row=row, column=3).value = f"VIR{random.randint(100000, 999999)}"
        ws.cell(row=row, column=5).value = f"{monthly_salary:.2f}"
        ws.cell(row=row, column=6).value = f"{balance:.2f}"
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    row += 1
    ws.cell(row=row, column=1).value = "Total (3 mois):"
    ws.cell(row=row, column=2).value = f"{monthly_salary * 3:.2f} MAD"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15
    
    wb.save(filename)
    print(f"✅ {filename}: Good (8,500 MAD x 3 months)")


def create_fair_income_xlsx(filename="fair_income_inconsistent.xlsx"):
    """Low salary, inconsistent - Score: 55-70"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"
    
    header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = "ATTIJARIWAFA BANK - FAIR PROFILE"
    ws['A1'].font = Font(size=16, bold=True, color="FF8C00")
    ws.merge_cells('A1:F1')
    ws['A2'] = "Client: Hassan Tazi - Freelancer"
    ws['A3'] = "Période: 3 mois (revenus irréguliers)"
    ws['A4'] = ""
    
    headers = ['Date', 'Description', 'Référence', 'Débit (MAD)', 'Crédit (MAD)', 'Solde (MAD)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 6
    balance = 3000.00
    incomes = [4500.00, 3200.00, 5800.00]  # Inconsistent
    
    for month, income in enumerate(incomes):
        income_date = datetime.now() - timedelta(days=75 - month*25)
        balance += income
        
        ws.cell(row=row, column=1).value = income_date.strftime('%d/%m/%Y')
        ws.cell(row=row, column=2).value = f"Virement - Projet {month+1}"
        ws.cell(row=row, column=3).value = f"VIR{random.randint(100000, 999999)}"
        ws.cell(row=row, column=5).value = f"{income:.2f}"
        ws.cell(row=row, column=6).value = f"{balance:.2f}"
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    row += 1
    ws.cell(row=row, column=1).value = "Total (irrégulier):"
    ws.cell(row=row, column=2).value = f"{sum(incomes):.2f} MAD"
    ws.cell(row=row, column=2).font = Font(bold=True, color="FF8C00")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15
    
    wb.save(filename)
    print(f"✅ {filename}: Fair (Inconsistent: 4,500-5,800 MAD)")


def create_excellent_loan_xlsx(filename="excellent_loan_history.xlsx"):
    """Perfect payment history - Score: 95-100"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan History"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = "CIH BANK - EXCELLENT CREDIT HISTORY"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws.merge_cells('A1:F1')
    ws['A2'] = "Client: Fatima Zahra"
    ws['A3'] = "Crédit Immobilier: 600,000 MAD @ 3.95%"
    ws['A4'] = "36 derniers paiements - TOUS À TEMPS"
    ws['A5'] = ""
    
    headers = ['Mois', 'Date', 'Mensualité', 'Statut', 'Retard', 'Solde Restant']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 7
    monthly_payment = 3500.00
    remaining = 550000.00
    
    for month in range(36):
        payment_date = datetime.now() - timedelta(days=1080 - month*30)
        remaining -= (monthly_payment * 0.7)
        
        ws.cell(row=row, column=1).value = f"Mois {month+1}"
        ws.cell(row=row, column=2).value = payment_date.strftime('%d/%m/%Y')
        ws.cell(row=row, column=3).value = f"{monthly_payment:.2f} MAD"
        ws.cell(row=row, column=4).value = "✓ Payé à temps"
        ws.cell(row=row, column=5).value = "0 jour"
        ws.cell(row=row, column=6).value = f"{remaining:.2f} MAD"
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        
        row += 1
    
    row += 1
    ws.cell(row=row, column=1).value = "BILAN: 36/36 paiements à temps - EXCELLENT"
    ws.cell(row=row, column=1).font = Font(bold=True, color="008000")
    ws.merge_cells(f'A{row}:F{row}')
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18
    
    wb.save(filename)
    print(f"✅ {filename}: Excellent (36/36 payments on time)")


def create_poor_loan_xlsx(filename="poor_loan_history.xlsx"):
    """Multiple late payments - Score: 40-60"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan History"
    
    header_fill = PatternFill(start_color="C82333", end_color="C82333", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws['A1'] = "WAFA BANK - POOR CREDIT HISTORY"
    ws['A1'].font = Font(size=16, bold=True, color="C82333")
    ws.merge_cells('A1:F1')
    ws['A2'] = "Client: Rachid Bennani"
    ws['A3'] = "Crédit Auto: 120,000 MAD @ 6.5%"
    ws['A4'] = "⚠️ MULTIPLE RETARDS DE PAIEMENT"
    ws['A5'] = ""
    
    headers = ['Mois', 'Date', 'Mensualité', 'Statut', 'Retard', 'Solde Restant']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row = 7
    monthly_payment = 2200.00
    remaining = 95000.00
    
    # Create pattern of late payments
    statuses = [
        ("Payé à temps", 0, "D4EDDA"),
        ("Retard 15j", 15, "FFF3CD"),
        ("Retard 30j", 30, "F8D7DA"),
        ("Payé à temps", 0, "D4EDDA"),
        ("Retard 7j", 7, "FFF3CD"),
        ("Retard 45j", 45, "F8D7DA"),
        ("Retard 10j", 10, "FFF3CD"),
        ("Payé à temps", 0, "D4EDDA"),
        ("Retard 20j", 20, "F8D7DA"),
        ("Retard 5j", 5, "FFF3CD"),
        ("Retard 60j", 60, "F8D7DA"),
        ("Payé à temps", 0, "D4EDDA"),
    ]
    
    for month, (status, delay, color) in enumerate(statuses):
        payment_date = datetime.now() - timedelta(days=360 - month*30)
        remaining -= (monthly_payment * 0.6)
        
        ws.cell(row=row, column=1).value = f"Mois {month+1}"
        ws.cell(row=row, column=2).value = payment_date.strftime('%d/%m/%Y')
        ws.cell(row=row, column=3).value = f"{monthly_payment:.2f} MAD"
        ws.cell(row=row, column=4).value = status
        ws.cell(row=row, column=5).value = f"{delay} jours"
        ws.cell(row=row, column=6).value = f"{remaining:.2f} MAD"
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        row += 1
    
    row += 1
    ws.cell(row=row, column=1).value = "BILAN: 8 retards sur 12 mois - MAUVAIS HISTORIQUE"
    ws.cell(row=row, column=1).font = Font(bold=True, color="C82333")
    ws.merge_cells(f'A{row}:F{row}')
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18
    
    wb.save(filename)
    print(f"✅ {filename}: Poor (8 late payments in 12 months)")


# ===== PDF FILES =====

def create_excellent_payslip_pdf(filename="payslip_excellent_senior.pdf"):
    """High salary payslip - Score: 95-100"""
    doc = SimpleDocTemplate(filename, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], 
                                 fontSize=18, textColor=colors.HexColor('#1e3a8a'), 
                                 alignment=TA_CENTER)
    
    story.append(Paragraph("MAROC TELECOM - Bulletin de Paie", title_style))
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph("<b>Employé:</b> Fatima Zahra Benali - Senior Manager", styles['Normal']))
    story.append(Paragraph(f"<b>Période:</b> Décembre 2025", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Salary breakdown
    data = [
        ['Élément', 'Montant (MAD)'],
        ['Salaire de Base', '15,000.00'],
        ['Prime d\'Ancienneté', '2,000.00'],
        ['Prime de Responsabilité', '3,000.00'],
        ['<b>Salaire Brut</b>', '<b>20,000.00</b>'],
        ['CNSS (4.48%)', '-896.00'],
        ['IR (38%)', '-7,600.00'],
        ['<b>Salaire Net</b>', '<b>11,504.00</b>'],
    ]
    
    table = Table(data, colWidths=[3.5*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#e7f3e7')),
        ('BACKGROUND', (0, 7), (-1, 7), colors.HexColor('#d4edda')),
        ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
        ('FONTNAME', (0, 7), (-1, 7), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(table)
    doc.build(story)
    print(f"✅ {filename}: Excellent (20,000 MAD gross, 11,504 net)")


def create_fair_payslip_pdf(filename="payslip_fair_entry_level.pdf"):
    """Low salary payslip - Score: 55-70"""
    doc = SimpleDocTemplate(filename, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], 
                                 fontSize=18, textColor=colors.HexColor('#FF8C00'), 
                                 alignment=TA_CENTER)
    
    story.append(Paragraph("COMPTOIR COMMERCIAL - Bulletin de Paie", title_style))
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph("<b>Employé:</b> Hassan Tazi - Vendeur", styles['Normal']))
    story.append(Paragraph(f"<b>Période:</b> Novembre 2025", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    data = [
        ['Élément', 'Montant (MAD)'],
        ['Salaire de Base', '3,500.00'],
        ['<b>Salaire Brut</b>', '<b>3,500.00</b>'],
        ['CNSS (4.48%)', '-156.80'],
        ['IR (10%)', '-350.00'],
        ['<b>Salaire Net</b>', '<b>2,993.20</b>'],
    ]
    
    table = Table(data, colWidths=[3.5*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF8C00')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(table)
    doc.build(story)
    print(f"✅ {filename}: Fair (3,500 MAD gross, 2,993 net)")


if __name__ == "__main__":
    print("📊 Generating VARIED test files for different credit scores...\n")
    print("=" * 60)
    
    print("\n📈 EXCELLENT PROFILE (Score: 90-100)")
    print("-" * 60)
    create_excellent_income_xlsx()
    create_excellent_loan_xlsx()
    create_excellent_payslip_pdf()
    
    print("\n📊 GOOD PROFILE (Score: 75-85)")
    print("-" * 60)
    create_good_income_xlsx()
    
    print("\n📉 FAIR PROFILE (Score: 55-70)")
    print("-" * 60)
    create_fair_income_xlsx()
    create_fair_payslip_pdf()
    
    print("\n❌ POOR PROFILE (Score: 40-60)")
    print("-" * 60)
    create_poor_loan_xlsx()
    
    print("\n" + "=" * 60)
    print("✅ ALL VARIED TEST FILES GENERATED!")
    print("=" * 60)
    print("\n📁 Files Created:")
    print("  EXCELLENT (90-100 score):")
    print("    • excellent_income_consistency.xlsx (18K MAD x 6 months)")
    print("    • excellent_loan_history.xlsx (36/36 on-time)")
    print("    • payslip_excellent_senior.pdf (20K MAD)")
    print("\n  GOOD (75-85 score):")
    print("    • good_income_consistency.xlsx (8.5K MAD x 3 months)")
    print("\n  FAIR (55-70 score):")
    print("    • fair_income_inconsistent.xlsx (Irregular: 3-6K MAD)")
    print("    • payslip_fair_entry_level.pdf (3.5K MAD)")
    print("\n  POOR (40-60 score):")
    print("    • poor_loan_history.xlsx (8 late payments)")
    print("\n💡 Use these files to test different credit scores!")
    print("💡 Mix and match to create various risk profiles")
